import tkinter as tk
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os

# Excel路径
file_path = os.path.join(os.path.dirname(__file__), "schedule.xlsx")

# 读取数据（去除空行）
df = pd.read_excel(file_path)
df = df[df['时间'].notna()].reset_index(drop=True)

# 加载 Excel 表格以写入
wb = load_workbook(file_path)
ws = wb.active
header = [cell.value for cell in ws[1]]
col_index = header.index("完成次数") + 1  # openpyxl 列从1开始

# 判断当前在哪一行
def get_current_active_row(df):
    now = datetime.now().time()
    for idx, row in df.iterrows():
        time_str = str(row['时间'])
        if '-' in time_str:
            try:
                start_str, end_str = time_str.split('-')
                start = datetime.strptime(start_str.strip(), "%H:%M").time()
                end = datetime.strptime(end_str.strip(), "%H:%M").time()
                if start <= now <= end:
                    return idx
            except:
                continue
    return -1

# 完美达成次数
def get_perfect_count():
    perfect = 0
    min=999
    for i in range(len(df)):
        count = df.at[i, '完成次数']
        if count <= min:
            min = count
    perfect = min
    return perfect
# 更新完成次数
def update_count(row_idx, delta):
    excel_row = row_idx + 2  # Excel从第2行是数据
    cell = ws.cell(row=excel_row, column=col_index)
    old_val = cell.value or 0
    new_val = max(0, int(old_val) + delta)
    cell.value = new_val
    df.at[row_idx, '完成次数'] = new_val
    wb.save(file_path)
    update_ui()

# 界面更新
def update_ui():
    for widget in frame.winfo_children():
        widget.destroy()

    # 表头
    tk.Label(frame, text="时间", width=15, bg="lightgray").grid(row=0, column=0)
    tk.Label(frame, text="事件", width=30, bg="lightgray").grid(row=0, column=1)
    tk.Label(frame, text="完成次数", width=10, bg="lightgray").grid(row=0, column=2)

    current_row = get_current_active_row(df)

    for i, row in df.iterrows():
        bg_color = "red" if i == current_row else None

        tk.Label(frame, text=row['时间'], width=15, bg=bg_color).grid(row=i + 1, column=0)
        tk.Label(frame, text=row['事件'], width=30, anchor="w", bg=bg_color).grid(row=i + 1, column=1)

        count_var = tk.StringVar(value=str(row['完成次数']))
        tk.Label(frame, textvariable=count_var, width=10, bg=bg_color).grid(row=i + 1, column=2)

        tk.Button(frame, text="+1", command=lambda i=i: update_count(i, 1)).grid(row=i + 1, column=3)
        tk.Button(frame, text="-1", command=lambda i=i: update_count(i, -1)).grid(row=i + 1, column=4)

    # 完美达成统计
    perfect_count = get_perfect_count()
    tk.Label(frame, text="完美达成次数", fg="red", bg="yellow", width=20).grid(row=len(df) + 1, column=0, columnspan=2)
    tk.Label(frame, text=str(perfect_count), fg="red", bg="yellow", width=10).grid(row=len(df) + 1, column=2)

# 初始化窗口
root = tk.Tk()
root.title("课程进度可视化")

# 滚动框架
canvas = tk.Canvas(root)
frame = tk.Frame(canvas)
scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)

scrollbar.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)
canvas.create_window((0, 0), window=frame, anchor='nw')
frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

update_ui()
root.mainloop()
